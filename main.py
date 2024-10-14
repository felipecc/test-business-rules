import csv
import json
import sqlite3
from datetime import datetime
from typing import Any, Dict, List, Union, cast
from database import get_oportunidades_dcaf, get_recebimento_benner, get_dados_provisionamento, get_dados_recebimento, read_query_from_file
from openpyxl import Workbook
from openpyxl.styles import Font
import polars as pl

import toolz
from business_rules import run_all
from business_rules.actions import BaseActions, rule_action
from business_rules.fields import FIELD_TEXT
from business_rules.variables import BaseVariables, numeric_rule_variable, string_rule_variable

from model import Action, AllConditions, Condition, DiasTolerancia, ListaCargoCarencia, Oportunidade, OportunidadeNaoProcessada, Recebimento, Rule
from utils import create_name_rule, extract_name_rule, extract_numbers_from_text, is_valid_date, is_valid_json, parse_datetime


decaf_benner = {
    "Manutenção": ["MANUTENÇÃO", "MANUTENÇÃO - MERCADO EXTERNO"],
    "Licença de Uso": ["LICENÇAS DE USO", "LICENÇAS DE USO - MERCADO EXTERNO"],
    "Consultoria": [
        "PRESTAÇÃO DE SERVIÇOS MERCADO INTERNO",
        "CONSULTORIA E IMPLANTAÇÃO",
        "CONSULTORIA E IMPLANTAÇÃO - MERCADO EXTERNO"
    ],
    "Subscrição": ["SUBSCRIÇÃO - MERCADO EXTERNO"],
    "Serviços": ["SERVIÇOS TECNICOS - MERCADO EXTERNO","SERVIÇOS SISTEMAS"],
    "Customização": ["SERVIÇOS SISTEMAS"],
    "Locação": ["LOCAÇÃO DE SISTEMAS"],
    "Outsourcing": ["OUTSOURCING"],
    "SAAS": ["SUBSCRIÇÃO"]
}


def mount_rule(
    *,
    name_rule: str,
    json_function_filter: str,
    validity_period_start: datetime,
    validity_period_end: datetime | None,
) -> Rule:
    function_filters = parse_json_filter(json_function_filter)
    function_name = extract_function_name(function_filters)
    filters = extract_filters(function_filters)

    list_conditions_instance = create_conditions(filters)
    add_validity_period_conditions(
        list_conditions_instance, validity_period_start, validity_period_end)

    all_conditions_instance = AllConditions(all=list_conditions_instance)
    list_actions_instance = [
        Action(name=function_name, params={"name_rule": name_rule})]

    return Rule(conditions=all_conditions_instance, actions=list_actions_instance)


def parse_json_filter(json_function_filter: str) -> Dict[str, Union[str, Any]]:
    if not is_valid_json(json_function_filter):
        raise ValueError("Invalid JSON format for function filter")
    return json.loads(json_function_filter)


def extract_function_name(function_filters: Dict) -> str:
    function_name = toolz.get("function_name", function_filters)
    if not function_name:
        raise ValueError("Missing function_name in function filters")
    return function_name


def extract_filters(function_filters: Dict) -> List[Dict]:
    filters = toolz.get("filter", function_filters)
    if not filters:
        raise ValueError("Missing filters in function filters")
    return filters


def create_conditions(filters: List[Dict]) -> List[Condition]:
    return [Condition(**condition) for condition in filters]


def add_validity_period_conditions(
    conditions: List[Condition],
    start_date: datetime,
    end_date: datetime | None
) -> None:
    start_timestamp = start_date.replace(
        hour=0, minute=1, second=0).timestamp()
    end_timestamp = (end_date.replace(hour=23, minute=59, second=59)
                     if end_date else datetime.now()).timestamp()

    conditions.extend([
        Condition(name="dt_prosposta_ganha",
                  operator="greater_than_or_equal_to", value=start_timestamp),
        Condition(name="dt_prosposta_ganha",
                  operator="less_than_or_equal_to", value=end_timestamp)
    ])

def get_nearest_day(days: List[int], day: int) -> int:
    return min(days, key=lambda x: (abs(x - day), x))


class OportunidadeVariables(BaseVariables):  # type: ignore
    def __init__(self, oportunidade: Oportunidade) -> None:
        self.__oportunidade = oportunidade

    @string_rule_variable(label="vertical")  # type: ignore
    def vertical(self) -> str:
        return self.__oportunidade.nm_vertical_item

    @string_rule_variable(label="contrato")  # type: ignore
    def contrato(self) -> str:
        return self.__oportunidade.ds_classificacao

    @string_rule_variable(label="evento")  # type: ignore
    def evento(self) -> str:
        return self.__oportunidade.tp_servico

    @numeric_rule_variable(label="data_proposta_ganha")  # type: ignore
    def dt_prosposta_ganha(self) -> float:
        return self.__oportunidade.dt_fechamento.timestamp()

    def match_data(self) -> bool:
        return True


class OportunidadeActions(BaseActions):  # type: ignore
    def __init__(self, oportunidade: Oportunidade, conn: sqlite3.Connection, workbook: Workbook) -> None:
        self.__oportunidade = oportunidade
        self.__conn = conn
        self.__workbook = workbook
        self.__ws_processados = workbook["Processados"]
        self.__ws_nao_processados = workbook["Não Processados"]

    def _get_tolerancia_dias_execucao_vencimento(
        self, dias_tolerancia: Dict[int, int], dias: int
    ) -> int | None:
        resultado = toolz.get_in([str(dias)], dias_tolerancia)

        return resultado[0] if resultado else None

    @rule_action(
        label="Funcao para computar o valor percentual distribuido entre os cargos",
        params={"name_rule": FIELD_TEXT},
    )  # type: ignore
    def calculo_percentual_cargo(self, name_rule: str) -> None:
        cargos_processados_vlr_e_percentuais = []
        registros_nao_processados = []

        id, nome_regra = extract_name_rule(name_rule)

        assert id, "id for rule need to be extracted"

        cursor = self.__conn.cursor()
        cursor.execute("SELECT * FROM regra WHERE id = ?", (id,))
        row = cursor.fetchone()

        if not row:
            raise Exception("Row with id={id} not found")

        dias_tolerancia_json: str = row[9]

        dias_tolerancia: Union[DiasTolerancia | None] = None

        if dias_tolerancia_json:
            if not (is_valid_json(dias_tolerancia_json)):
                raise Exception(
                    "Error for parsing json file dias_tolerancia_json")

            dias_tolerancia = DiasTolerancia.model_validate_json(
                json.loads(dias_tolerancia_json)
            )

        cargo_carencia_json: str = row[6]

        if cargo_carencia_json:
            if not (is_valid_json(cargo_carencia_json)):
                raise Exception(
                    "Error for parsing json file in column cargo_carencia_json")

        cargos_carencia: ListaCargoCarencia = ListaCargoCarencia.model_validate_json(
            cargo_carencia_json
        )

        vlr_item = float(self.__oportunidade.nr_preco_total_desconto)

        primeira_parcela_dias = (
            extract_numbers_from_text(self.__oportunidade.ds_primeira_parcela)
            if self.__oportunidade.ds_primeira_parcela
            else None
        )
        
        group_receita_master = decaf_benner.get(self.__oportunidade.tp_servico, None)
        
        #Todo avaliar uma forma melhor de agrupar estes labels
        if not len(group_receita_master) > 0:
            raise Exception(f'Oportunidade: {self.__oportunidade.oportunidade_id} sem um grupo depara do dcaf x benner')
        
        recebimentos: List[Recebimento] = []
        for group in group_receita_master:
            registro = get_recebimento_benner(nr_pedido_fq=self.__oportunidade.ds_contrato_benner, receita_master=group)
            if registro:
                recebimentos.extend(registro)
        
        if not recebimentos:
            return None
        
        for recebimento in recebimentos:
            if primeira_parcela_dias:
                if len(cargos_carencia.cargos) > 1:
                    for cargo in cargos_carencia.cargos:

                        range_days = [int(k) for k in cargo.carencia.keys()]
                        nearest_day = get_nearest_day(
                            range_days, int(primeira_parcela_dias))

                        processado = {
                            "oportunidade": self.__oportunidade.oportunidade_id,
                            "nm_vertical_item": self.__oportunidade.nm_vertical_item,
                            "regra_nome": nome_regra,
                            "cargo": cargo.cargo,
                            "vlr_percentual": cargo.carencia.get(
                                str(nearest_day)
                            ),
                            "vlr_comissao": cast(
                                float, cargo.carencia.get(
                                    str(nearest_day))
                            )
                            * recebimento.valor_liquido ,
                        }
                        cargos_processados_vlr_e_percentuais.append(processado)

                else:
                    cargo = toolz.first(cargos_carencia.cargos)
                    processado = {
                        "cargo": cargo.cargo,
                        "vlr_percentual": cargo.carencia.get("0"),
                        "vlr_comissao": cast(float, cargo.carencia.get("0")) * vlr_item,
                    }
                    cargos_processados_vlr_e_percentuais.append(processado)
            else:
                oportunidade_nao_processada: OportunidadeNaoProcessada = (
                    OportunidadeNaoProcessada(
                        oportunidade=self.__oportunidade,
                        motivo="valor no campo primeira parcela vazio",
                        regra=f"{id} - {nome_regra}",
                    )
                )
                registros_nao_processados.append(oportunidade_nao_processada)

        self._write_to_excel(
            cargos_processados_vlr_e_percentuais, registros_nao_processados)

    def _create_header(self, workbook: Workbook, headers: List[str]) -> None:
        workbook.append(headers)
        for cell in workbook[1]:
            cell.font = Font(bold=True)

    def _write_to_excel(self, cargos_processados: List[Dict[str, Any]], registros_nao_processados: List[OportunidadeNaoProcessada]) -> None:
        SHEET_NAME_PROCESSED = "Processados"
        SHEET_NAME_UNPROCESSED = "Não Processados"
        HEADERS_PROCESSED = ["Oportunidade ID", "Vertical Item",
                             "Regra Nome", "Cargo", "Valor Percentual", "Valor Comissão"]
        HEADERS_UNPROCESSED = ["Oportunidade ID", "Motivo", "Regra"]

        if self.__ws_processados is None:
            self.__ws_processados = self.__workbook.create_sheet(
                SHEET_NAME_PROCESSED)
            self._create_header(self.__ws_processados, HEADERS_PROCESSED)

        if self.__ws_nao_processados is None:
            self.__ws_nao_processados = self.__workbook.create_sheet(
                SHEET_NAME_UNPROCESSED)
            self._create_header(self.__ws_nao_processados, HEADERS_UNPROCESSED)

        # Write data for processados
        processados_data = [
            [
                item["oportunidade"],
                item["nm_vertical_item"],
                item["regra_nome"],
                item["cargo"],
                item["vlr_percentual"],
                item["vlr_comissao"]
            ] for item in cargos_processados
        ]
        for row in processados_data:
            self.__ws_processados.append(row)

        # Write data for não processados
        nao_processados_data = [
            [
                item.oportunidade.oportunidade_id,
                item.motivo,
                item.regra
            ] for item in registros_nao_processados
        ]
        for row in nao_processados_data:
            self.__ws_nao_processados.append(row)


if __name__ == "__main__":
    
    rs_provisionamentos = get_dados_provisionamento()
    provisionamentos = [r.model_dump() for r in rs_provisionamentos]
    df_provisionamentos = pl.DataFrame(provisionamentos)
    
    rs_recebimentos_detalhados = get_dados_recebimento()
    recebimentos_detalhados = [r.model_dump() for r in rs_recebimentos_detalhados]
    df_recebimentos_detalhados = pl.DataFrame(recebimentos_detalhados)
    
    with pl.SQLContext(
        provisionamentos=df_provisionamentos,
        recebimentos=df_recebimentos_detalhados,
        eager=True
    ) as ctx:
        query = read_query_from_file('query_df.sql')
        df = ctx.execute(query)
        df.write_csv('output.csv')    




    # ID = 0
    # NOME = 1
    # TIPO = 2
    # DATA_CRIACAO = 3
    # DATA_PUBLICACAO = 4
    # FUNCAO_FILTRO = 5
    # CARGO_CARENCIA = 6
    # INICIO_VIGENCIA = 7
    # FIM_VIGENCIA = 8
    # DIAS_TOLERANCIA = 9
    # conn = sqlite3.connect("example.db")
    # cursor = conn.cursor()

    # workbook = Workbook()
    # workbook.remove(workbook.active)  # Remove the default sheet

    # ws_processados = workbook.create_sheet("Processados")
    # ws_nao_processados = workbook.create_sheet("Não Processados")

    # headers_processados = ["Oportunidade ID", "Vertical Item",
    #                        "Regra Nome", "Cargo", "Valor Percentual", "Valor Comissão"]
    # ws_processados.append(headers_processados)
    # for cell in ws_processados[1]:
    #     cell.font = Font(bold=True)

    # headers_nao_processados = ["Oportunidade ID", "Motivo", "Regra"]
    # ws_nao_processados.append(headers_nao_processados)
    # for cell in ws_nao_processados[1]:
    #     cell.font = Font(bold=True)

    # oportunidades = get_oportunidades_dcaf('01/01/2023', '01/01/2023')

    # cursor.execute("SELECT * FROM regra")
    # rows = cursor.fetchall()

    # mounted_rules: List[Rule] = []

    # for row in rows:
    #     name_rule = create_name_rule(row[ID], row[NOME])

    #     assert name_rule, "name_rule is not could be empty. "

    #     function_filters = row[FUNCAO_FILTRO]
    #     validity_period_start = (
    #         datetime.strptime(
    #             row[INICIO_VIGENCIA], "%Y-%m-%d") if is_valid_date(row[INICIO_VIGENCIA]) else None
    #     )

    #     assert validity_period_start, "validity_period_start is not could be empty. "

    #     validity_period_end = (
    #         datetime.strptime(
    #             row[FIM_VIGENCIA], "%Y-%m-%d") if is_valid_date(row[FIM_VIGENCIA]) else None
    #     )

    #     mounted_rules.append(mount_rule(
    #         name_rule=name_rule,
    #         json_function_filter=function_filters,
    #         validity_period_start=validity_period_start,
    #         validity_period_end=validity_period_end,
    #     ))

    # for oportunidade in oportunidades:
    #     run_all(
    #         rule_list=[rule.model_dump() for rule in mounted_rules],
    #         defined_variables=OportunidadeVariables(oportunidade),
    #         defined_actions=OportunidadeActions(oportunidade, conn, workbook),
    #         stop_on_first_trigger=True,
    #     )

    # conn.close()
    # workbook.save("oportunidades_processadas.xlsx")
